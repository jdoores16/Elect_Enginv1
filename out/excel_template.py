
from pathlib import Path
from typing import Optional, Dict, List, Tuple
import logging
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Fill, Alignment, Border

logger = logging.getLogger(__name__)


def find_template(bucket_dir: Path, session_prefix: str = "") -> Optional[Path]:
    """
    Find an Excel template file in the bucket directory.
    Looks for files with 'template' in the name or ending in '_template.xlsx'
    Only returns .xlsx or .xlsm files that openpyxl can read.
    """
    if not bucket_dir.exists():
        return None
    
    candidates = []
    for p in bucket_dir.iterdir():
        if not p.is_file():
            continue
        # Only accept formats that openpyxl can load
        if p.suffix.lower() not in ['.xlsx', '.xlsm']:
            continue
        # Check if it matches session
        if session_prefix and not p.name.startswith(session_prefix):
            continue
        # Check if it looks like a template
        name_lower = p.name.lower()
        if 'template' in name_lower or '_tmpl' in name_lower:
            candidates.append(p)
    
    if candidates:
        # Return the most recent template
        candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        logger.info(f"Found template: {candidates[0].name}")
        return candidates[0]
    
    logger.info("No template file found in bucket")
    return None


def read_template_structure(template_path: Path) -> Dict:
    """
    Read the structure of an Excel template.
    Returns a dict with column headers, formatting info, and sheet structure.
    """
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Read header row (assume it's row 1)
        headers = []
        header_styles = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
                # Store cell formatting
                header_styles.append({
                    'font': cell.font.copy() if cell.font else None,
                    'fill': cell.fill.copy() if cell.fill else None,
                    'alignment': cell.alignment.copy() if cell.alignment else None,
                    'border': cell.border.copy() if cell.border else None
                })
            else:
                break
        
        # Get column widths
        column_widths = {}
        for i, col in enumerate(ws.columns, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            if ws.column_dimensions[col_letter].width:
                column_widths[i] = ws.column_dimensions[col_letter].width
        
        structure = {
            'headers': headers,
            'header_styles': header_styles,
            'column_widths': column_widths,
            'sheet_name': ws.title
        }
        
        logger.info(f"Template structure: {len(headers)} columns - {headers}")
        return structure
        
    except Exception as e:
        logger.error(f"Error reading template structure: {e.__class__.__name__}: {e}")
        raise


def apply_template_to_data(
    circuits: List[Tuple[str, str]], 
    panel_name: str,
    template_path: Optional[Path],
    output_path: Path
) -> Path:
    """
    Create an Excel file using template formatting, filled with OCR circuit data.
    If no template is provided or template fails to load, creates a basic formatted schedule.
    """
    if template_path and template_path.exists():
        try:
            logger.info(f"Using template: {template_path.name}")
            structure = read_template_structure(template_path)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = structure.get('sheet_name', 'Panel_Schedule')
            
            # Apply headers with formatting
            headers = structure.get('headers', ['Panel', 'Circuit', 'Description'])
            header_styles = structure.get('header_styles', [])
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                # Apply header formatting if available
                if col_idx - 1 < len(header_styles):
                    style = header_styles[col_idx - 1]
                    if style.get('font'):
                        cell.font = style['font']
                    if style.get('fill'):
                        cell.fill = style['fill']
                    if style.get('alignment'):
                        cell.alignment = style['alignment']
                    if style.get('border'):
                        cell.border = style['border']
            
            # Apply column widths
            for col_idx, width in structure.get('column_widths', {}).items():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width
            
            # Fill in circuit data
            # Map circuit data to template columns
            for row_idx, (circuit_num, description) in enumerate(circuits, 2):
                # Try to intelligently map to template columns
                for col_idx, header in enumerate(headers, 1):
                    header_lower = header.lower()
                    if 'panel' in header_lower or 'board' in header_lower:
                        ws.cell(row=row_idx, column=col_idx, value=panel_name)
                    elif 'circuit' in header_lower or 'ckt' in header_lower or 'number' in header_lower:
                        ws.cell(row=row_idx, column=col_idx, value=circuit_num)
                    elif 'desc' in header_lower or 'load' in header_lower or 'name' in header_lower:
                        ws.cell(row=row_idx, column=col_idx, value=description)
            
            logger.info(f"Applied template with {len(circuits)} circuits")
            
        except Exception as e:
            logger.warning(f"Failed to apply template {template_path.name}: {e.__class__.__name__}: {e}. Falling back to basic format.")
            # Fall through to basic format creation below
            template_path = None
    
    if not template_path:
        # Create basic formatted schedule
        logger.info("No template found, creating basic formatted schedule")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Panel_Schedule'
        
        # Create headers with basic formatting
        headers = ['Panel', 'Circuit', 'Description']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        
        # Fill in circuit data
        for row_idx, (circuit_num, description) in enumerate(circuits, 2):
            ws.cell(row=row_idx, column=1, value=panel_name)
            ws.cell(row=row_idx, column=2, value=circuit_num)
            ws.cell(row=row_idx, column=3, value=description)
    
    wb.save(output_path)
    logger.info(f"Saved panel schedule to {output_path}")
    return output_path
