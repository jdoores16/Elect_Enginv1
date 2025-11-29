"""
PDF Export Module

Converts Excel spreadsheets to PDF using LibreOffice for exact copy.
Falls back to reportlab-based generation if LibreOffice fails.
"""
from __future__ import annotations
import subprocess
import logging
import shutil
import tempfile
from pathlib import Path
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from app.schemas.panel_ir import PanelScheduleIR

logger = logging.getLogger(__name__)


def convert_excel_to_pdf(excel_path: str, output_pdf: str) -> str:
    """
    Convert Excel file to PDF using LibreOffice.
    
    This produces an exact visual copy of the spreadsheet.
    
    Args:
        excel_path: Path to the Excel file
        output_pdf: Desired output PDF path
        
    Returns:
        Path to the generated PDF file
    """
    excel_path = Path(excel_path)
    output_pdf = Path(output_pdf)
    
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Create temp directory for LibreOffice output
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_dir = Path(temp_dir)
        
        try:
            # Run LibreOffice in headless mode to convert to PDF
            result = subprocess.run(
                [
                    'libreoffice',
                    '--headless',
                    '--convert-to', 'pdf',
                    '--outdir', str(temp_dir),
                    str(excel_path)
                ],
                capture_output=True,
                text=True,
                timeout=60  # 60 second timeout
            )
            
            if result.returncode != 0:
                logger.error(f"LibreOffice conversion failed: {result.stderr}")
                raise RuntimeError(f"LibreOffice conversion failed: {result.stderr}")
            
            # Find the generated PDF
            generated_pdf = temp_dir / excel_path.with_suffix('.pdf').name
            
            if not generated_pdf.exists():
                # Try to find any PDF in the temp directory
                pdfs = list(temp_dir.glob('*.pdf'))
                if pdfs:
                    generated_pdf = pdfs[0]
                else:
                    raise FileNotFoundError("LibreOffice did not generate a PDF file")
            
            # Move to final location
            shutil.copy2(generated_pdf, output_pdf)
            logger.info(f"Successfully converted {excel_path} to {output_pdf}")
            
            return str(output_pdf)
            
        except subprocess.TimeoutExpired:
            logger.error("LibreOffice conversion timed out")
            raise RuntimeError("PDF conversion timed out")
        except Exception as e:
            logger.error(f"LibreOffice conversion error: {e}")
            raise


def export_pdf_from_excel(excel_path: str, output_pdf: str = None) -> str:
    """
    Export PDF from an Excel file.
    
    Uses LibreOffice for exact copy. Falls back to basic text export if LibreOffice fails.
    
    Args:
        excel_path: Path to the Excel file
        output_pdf: Optional output path (defaults to same name with .pdf extension)
        
    Returns:
        Path to the generated PDF file
    """
    excel_path = Path(excel_path)
    
    if output_pdf is None:
        output_pdf = excel_path.with_suffix('.pdf')
    else:
        output_pdf = Path(output_pdf)
    
    try:
        return convert_excel_to_pdf(str(excel_path), str(output_pdf))
    except Exception as e:
        logger.warning(f"LibreOffice conversion failed, using fallback: {e}")
        # Fall back to basic text-based PDF (not ideal but better than nothing)
        return _fallback_excel_to_pdf(str(excel_path), str(output_pdf))


def _fallback_excel_to_pdf(excel_path: str, output_pdf: str) -> str:
    """
    Fallback PDF generation using openpyxl + reportlab.
    
    This doesn't preserve exact formatting but captures the data.
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    c = canvas.Canvas(output_pdf, pagesize=letter)
    width, height = letter
    
    c.setFont("Helvetica-Bold", 12)
    c.drawString(0.5*inch, height - 0.5*inch, f"Panel Schedule (Fallback Export)")
    
    c.setFont("Helvetica", 8)
    y = height - 0.8*inch
    
    # Export visible rows
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=True):
        if y < 0.5*inch:
            c.showPage()
            y = height - 0.5*inch
            c.setFont("Helvetica", 8)
        
        # Build row text
        row_text = "  ".join(str(cell) if cell is not None else "" for cell in row[:15])
        c.drawString(0.5*inch, y, row_text[:120])  # Truncate long rows
        y -= 10
    
    c.save()
    return output_pdf


def export_pdf_from_ir(ir: PanelScheduleIR, out_pdf: str) -> str:
    """
    Legacy function: Generate PDF from Panel IR using reportlab.
    
    NOTE: This is kept for backwards compatibility. For exact Excel copies,
    use export_pdf_from_excel() instead.
    """
    c = canvas.Canvas(out_pdf, pagesize=letter)
    width, height = letter

    # Title
    c.setFont("Helvetica-Bold", 14)
    c.drawString(0.75*inch, height - 0.9*inch, f"Panel Schedule: {ir.header.panel_name}")

    # Normalize phase label for text output
    for p in ir.header.left_params:
        if p.name_text.strip().upper() == "PHASE" and isinstance(p.value, str):
            p.value = p.value.replace("Ø", "").replace("O", "").replace("PHASE", "").strip().upper()
            if not p.value.endswith("PH"):
                p.value += "PH"

    # Header pairs (left / right columns)
    c.setFont("Helvetica", 9)
    y_left = height - 1.15*inch
    for pair in ir.header.left_params:
        val = "" if pair.value is None else str(pair.value)
        c.drawString(0.75*inch, y_left, f"{pair.name_text}: {val}")
        y_left -= 12

    y_right = height - 1.15*inch
    for pair in ir.header.right_params:
        val = "" if pair.value is None else str(pair.value)
        c.drawString(3.9*inch, y_right, f"{pair.name_text}: {val}")
        y_right -= 12

    y = min(y_left, y_right) - 16

    # Circuits header
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75*inch, y, "Circuits (summary)")
    y -= 14
    c.setFont("Helvetica", 8)
    c.drawString(0.75*inch, y, "CKT  SIDE  ROW  POLES  BRKR_A  LOAD_A  PH[A,B,C]  DESCRIPTION")
    y -= 10
    c.line(0.75*inch, y, width - 0.75*inch, y)
    y -= 8

    # Rows
    def phases(rec) -> str:
        return "".join(ch for ch, flag in zip("ABC", [rec.phA, rec.phB, rec.phC]) if flag)

    for rec in sorted(ir.circuits, key=lambda r: r.ckt):
        if y < 1.0*inch:
            c.showPage()
            y = height - 0.9*inch
            c.setFont("Helvetica", 8)

        phs = phases(rec)
        poles = "" if rec.poles is None else str(rec.poles)
        c.drawString(
            0.75*inch, y,
            f"{rec.ckt:<4} {rec.side:<4} {rec.excel_row:<3} {poles:<5} "
            f"{int(rec.breaker_amps):<6} {int(rec.load_amps):<6} {phs:<7} {rec.description or ''}"
        )
        y -= 10

    c.setFont("Helvetica-Oblique", 7)
    c.drawString(0.75*inch, 0.6*inch, "Draft – verify per current NEC and local AHJ before issuance.")
    c.save()
    return out_pdf
