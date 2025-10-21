from __future__ import annotations
from pathlib import Path
from typing import Optional, Dict
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from app.schemas.panel_ir import PanelScheduleIR

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = ROOT / "templates" / "panelboard_template.xlsx"  # master template

ODD_COLS  = {"desc": "A", "phaseA": "B", "phaseB": "C", "phaseC": "D", "pole": "E", "breaker": "F"}
EVEN_COLS = {"desc": "O", "phaseA": "L", "phaseB": "M", "phaseC": "N", "pole": "K", "breaker": "J"}

def _phase_slot_for_circuit(ckt: int) -> str:
    i = (ckt - 1) % 6
    if i in (0, 1): return "phaseA"
    if i in (2, 3): return "phaseB"
    return "phaseC"

def _top_left_of_range(addr: str) -> tuple[int, int]:
    c1, r1, _, _ = range_boundaries(addr)  # (min_col, min_row, max_col, max_row)
    return r1, c1

def _write_raw(ws, addr: str, value):
    """Write a raw value. For merged ranges, write only to the top-left cell."""
    if ":" in addr:
        r1, c1 = _top_left_of_range(addr)
        ws.cell(row=r1, column=c1).value = value
    else:
        ws[addr].value = value

def _get_param(ir: PanelScheduleIR, label: str) -> str:
    """Fetch a header parameter (by label)."""
    for p in ir.header.left_params:
        if p.name_text.strip().upper() == label.upper():
            return "" if p.value is None else str(p.value).strip()
    return ""

def _assert_template(ws):
    def normalize(s: str) -> str:
        s = (s or "").strip().upper()
        while s.endswith(":"):
            s = s[:-1].rstrip()
        return s
    sentinels = {
        "A2":"VOLTAGE","A3":"PHASE","A4":"WIRE","A5":"MAIN BUS AMPS","A6":"MAIN CIRCUIT BREAKER",
        "A7":"MOUNTING","A8":"FEED","A9":"FEED-THRU LUGS",
        "I2":"LOCATION","I3":"FED FROM","I4":"UL LISTED EQUIPMENT SHORT CIRCUIT RATING",
        "I5":"MAXIMUM AVAILABLE SHORT CIRCUIT CURRENT","I6":"PHASE CONDUCTOR",
        "I7":"NEUTRAL CONDUCTOR","I8":"GROUND CONDUCTOR",
    }
    for addr, exp in sentinels.items():
        got = normalize(ws[addr].value)
        if got != normalize(exp):
            raise ValueError(f"Template mismatch at {addr}: expected '{exp}', got '{ws[addr].value}'")

def _sanitize_filename(s: str) -> str:
    return "".join(ch if ch not in '\\/:*?"<>|' else "_" for ch in s).replace(" ", "_")

def _sanitize_sheet_title(s: str) -> str:
    invalid = set('[]:*?/\\')
    s = "".join("_" if ch in invalid else ch for ch in s).strip() or "SCHEDULE"
    return s[:31]

def write_excel_from_ir(
    ir: PanelScheduleIR,
    out_path: str,
    template_xlsx: Optional[str] = None,
    formulas: Optional[Dict[str, str]] = None,
):
    tpl = Path(template_xlsx) if template_xlsx else TEMPLATE_PATH
    if not tpl.exists():
        raise FileNotFoundError(f"Panelboard template not found: {tpl}")

    wb = load_workbook(tpl, data_only=False, keep_vba=False)
    ws = wb.active

    _assert_template(ws)

    # --- Header / labels ---
    _write_raw(ws, ir.header.panel_name_cell, ir.header.panel_name)
    for p in ir.header.left_params:
        _write_raw(ws, p.value_cell, p.value)
    for p in ir.header.right_params:
        _write_raw(ws, p.value_cell, p.value)
    _write_raw(ws, ir.header.right_unused_value_cell, None)

    # --- Circuits (rows 12..53) ---
    for c in ir.circuits:
        r = c.excel_row
        cols = ODD_COLS if c.side == "odd" else EVEN_COLS
        phase_slot = _phase_slot_for_circuit(c.ckt)

        if c.description is not None:
            ws[f"{cols['desc']}{r}"].value = c.description
        ws[f"{cols['breaker']}{r}"].value = float(c.breaker_amps)
        ws[f"{cols[phase_slot]}{r}"].value = float(c.load_amps)
        if c.poles is not None:
            ws[f"{cols['pole']}{r}"].value = int(c.poles)

    # --- KVA formulas (once) ---
    if formulas:
        for cell, formula in formulas.items():
            ws[cell].value = formula

    # --- File naming & sheet title ---
    file_voltage = _sanitize_filename(str(next(
        (p.value for p in ir.header.left_params if p.name_text.upper() == "VOLTAGE"),
        "UNKNOWN",
    )).strip())
    panel_name = _sanitize_filename(ir.header.panel_name.strip())
    out_p = Path(out_path).parent / f"panel_{panel_name}_{file_voltage}.xlsx"
    out_p.parent.mkdir(parents=True, exist_ok=True)

    tab_voltage = _get_param(ir, "VOLTAGE")
    tab_phase   = _get_param(ir, "PHASE")
    tab_wire    = _get_param(ir, "WIRE")
    ws.title = _sanitize_sheet_title(f"{tab_voltage}, {tab_phase}, {tab_wire}".strip(", "))

    wb.save(out_p)
    return out_p