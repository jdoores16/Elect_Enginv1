"""
Generate Panel Schedule Variable List Excel file.

Creates a three-column spreadsheet listing all panel schedule
variables, their current values, and confidence scores.
"""
import logging
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)


def generate_variable_list_excel(
    output_path: Path,
    panel_name: str,
    panel_specs: Dict[str, Any],
    circuits: Dict[str, Dict[str, Any]],
    confidence_data: Optional[Dict[str, Dict[str, Any]]] = None
) -> Path:
    """
    Generate a variable list Excel file with all panel schedule parameters.
    
    Args:
        output_path: Path to save the Excel file
        panel_name: Panel identifier
        panel_specs: Dictionary of panel specifications (voltage, phase, etc.)
        circuits: Dictionary of circuit data keyed by circuit number
        confidence_data: Optional dictionary with confidence info per parameter
                        Format: {param_name: {'confidence': float, 'method': str, 'source': str}}
        
    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Variable List"
    
    confidence_data = confidence_data or {}
    
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    border = Side(style='thin', color='000000')
    cell_border = Border(left=border, right=border, top=border, bottom=border)
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    
    ws['A1'] = "Variable Name"
    ws['B1'] = "Value"
    ws['C1'] = "Confidence"
    ws['A1'].font = header_font_white
    ws['B1'].font = header_font_white
    ws['C1'].font = header_font_white
    ws['A1'].fill = header_fill
    ws['B1'].fill = header_fill
    ws['C1'].fill = header_fill
    ws['A1'].border = cell_border
    ws['B1'].border = cell_border
    ws['C1'].border = cell_border
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['C1'].alignment = Alignment(horizontal='center')
    
    row = 2
    
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    high_conf_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    med_conf_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_conf_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def get_confidence_fill(confidence: float) -> PatternFill:
        """Return color based on confidence level."""
        if confidence >= 0.8:
            return high_conf_fill
        elif confidence >= 0.5:
            return med_conf_fill
        else:
            return low_conf_fill
    
    def add_section_header(title: str):
        nonlocal row
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].border = cell_border
        ws[f'B{row}'].border = cell_border
        ws[f'C{row}'].border = cell_border
        row += 1
    
    def add_variable(name: str, value: Any, param_key: Optional[str] = None):
        nonlocal row
        ws[f'A{row}'] = name
        ws[f'B{row}'] = str(value) if value is not None else ""
        ws[f'A{row}'].border = cell_border
        ws[f'B{row}'].border = cell_border
        ws[f'C{row}'].border = cell_border
        ws[f'A{row}'].alignment = Alignment(horizontal='left')
        ws[f'B{row}'].alignment = Alignment(horizontal='left')
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        
        if param_key and param_key in confidence_data:
            conf = confidence_data[param_key].get('effective_confidence', 0)
            ws[f'C{row}'] = f"{conf:.0%}"
            ws[f'C{row}'].fill = get_confidence_fill(conf)
        else:
            ws[f'C{row}'] = ""
        
        row += 1
    
    add_section_header("PANEL IDENTIFICATION")
    add_variable("Panel Name", panel_name, "panel_name")
    
    add_section_header("PANEL SPECIFICATIONS")
    add_variable("Voltage", panel_specs.get("voltage", ""), "voltage")
    add_variable("Phase", panel_specs.get("phase", ""), "phase")
    add_variable("Wire", panel_specs.get("wire", ""), "wire")
    add_variable("Main Bus Amps", panel_specs.get("main_bus_amps", ""), "main_bus_amps")
    add_variable("Main Circuit Breaker", panel_specs.get("main_breaker", ""), "main_breaker")
    add_variable("Mounting", panel_specs.get("mounting", ""), "mounting")
    add_variable("Feed", panel_specs.get("feed", ""), "feed")
    add_variable("Location", panel_specs.get("location", ""), "location")
    add_variable("Fed From", panel_specs.get("fed_from", ""), "fed_from")
    add_variable("Number of Circuits", panel_specs.get("number_of_ckts", len(circuits)), "number_of_ckts")
    
    if circuits:
        add_section_header("CIRCUIT DATA")
        
        sorted_circuits = sorted(circuits.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999)
        
        for circuit_num, circuit_data in sorted_circuits:
            if circuit_data.get('is_continuation'):
                continue
                
            desc = circuit_data.get('description', '')
            breaker_amps = circuit_data.get('breaker_amps', 0)
            poles = circuit_data.get('poles', 1)
            load_amps = circuit_data.get('load_amps', 0)
            load_type = circuit_data.get('load_type', '')
            circuit_conf = circuit_data.get('confidence', 0)
            
            add_variable(f"Pole Space {circuit_num} Description", desc, f"circuit_{circuit_num}_description")
            add_variable(f"Pole Space {circuit_num} Breaker Amps", f"{breaker_amps}A" if breaker_amps else "", f"circuit_{circuit_num}_breaker")
            add_variable(f"Pole Space {circuit_num} Poles", f"{poles}P" if poles else "", f"circuit_{circuit_num}_poles")
            if load_amps:
                add_variable(f"Pole Space {circuit_num} Load Amps", f"{load_amps}A", f"circuit_{circuit_num}_load")
            if load_type:
                add_variable(f"Pole Space {circuit_num} Load Type", load_type, f"circuit_{circuit_num}_load_type")
    
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    logger.info(f"Generated variable list Excel: {output_path}")
    
    return output_path
