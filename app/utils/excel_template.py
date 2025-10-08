
from pathlib import Path
from typing import Optional, Dict, List, Tuple
import logging
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Fill, Alignment, Border

logger = logging.getLogger(__name__)


def find_template(bucket_dir: Path, session_prefix: str = "") -> Optional[Path]:
    """
    Find an Excel template file in the bucket directory or templates folder.
    Looks for files with 'template' in the name or ending in '_template.xlsx'
    Only returns .xlsx or .xlsm files that openpyxl can read.
    Falls back to default template in templates/ folder if no user template found.
    """
    candidates = []
    
    # First, check bucket directory for user-uploaded templates
    if bucket_dir.exists():
        for p in bucket_dir.iterdir():
            if not p.is_file():
                continue
            # Only accept formats that openpyxl can load
            if p.suffix.lower() not in ['.xlsx', '.xlsm']:
                continue
            # Check if it matches session
            if session_prefix and not p.name.startswith(session_prefix):
                continue
            # Check if it looks like a template
            name_lower = p.name.lower()
            if 'template' in name_lower or '_tmpl' in name_lower:
                candidates.append(p)
    
    if candidates:
        # Return the most recent user template
        candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        logger.info(f"Found user template: {candidates[0].name}")
        return candidates[0]
    
    # Fallback to default template in templates folder
    templates_dir = Path(__file__).parent.parent.parent / "templates"
    default_template = templates_dir / "default_panelboard_template.xlsx"
    
    if default_template.exists():
        logger.info(f"Using default template: {default_template.name}")
        return default_template
    
    logger.info("No template file found in bucket or templates folder")
    return None


def read_template_structure(template_path: Path) -> Dict:
    """
    Read the structure of an Excel template.
    Returns a dict with column headers, formatting info, and sheet structure.
    """
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Read header row (assume it's row 1)
        headers = []
        header_styles = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
                # Store cell formatting
                header_styles.append({
                    'font': cell.font.copy() if cell.font else None,
                    'fill': cell.fill.copy() if cell.fill else None,
                    'alignment': cell.alignment.copy() if cell.alignment else None,
                    'border': cell.border.copy() if cell.border else None
                })
            else:
                break
        
        # Get column widths
        column_widths = {}
        for i, col in enumerate(ws.columns, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            if ws.column_dimensions[col_letter].width:
                column_widths[i] = ws.column_dimensions[col_letter].width
        
        structure = {
            'headers': headers,
            'header_styles': header_styles,
            'column_widths': column_widths,
            'sheet_name': ws.title
        }
        
        logger.info(f"Template structure: {len(headers)} columns - {headers}")
        return structure
        
    except Exception as e:
        logger.error(f"Error reading template structure: {e.__class__.__name__}: {e}")
        raise


def apply_template_to_data(
    circuits: List[Tuple[str, str]], 
    panel_name: str,
    template_path: Optional[Path],
    output_path: Path,
    panel_specs: Optional[Dict] = None
) -> Path:
    """
    Create an Excel file using template formatting, filled with OCR circuit data.
    If template is provided, copies entire template and updates:
    - Row 1, Column 1 (A1): Panel identifier/name
    - Rows 2-9, Column B: Panel specifications (voltage, phase, etc.) based on labels in Column A
    - Circuit data starting from data table row
    If no template is provided or template fails to load, creates a basic formatted schedule.
    """
    if template_path and template_path.exists():
        try:
            logger.info(f"Using template: {template_path.name}")
            
            # Load the entire template workbook
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Update Row 1, Column 1 (A1) with panel name - this is the panel identifier
            # A2-A9 are field labels (not editable), B2-B9 are field values (editable)
            ws.cell(row=1, column=1, value=panel_name)
            logger.info(f"Updated panel identifier (A1) to: {panel_name}")
            
            # Update panel specifications based on template labels
            # A2-A9 labels -> B2-B9 values
            # N2-N9 labels -> O2-O9 values
            # If panel_specs is None, skip this section entirely
            if panel_specs is not None:
                # Mapping of common label keywords to spec keys
                label_mapping = {
                    'voltage': ['voltage', 'volt'],
                    'phase': ['phase', 'ph'],
                    'wire': ['wire', 'wires'],
                    'main_bus_amps': ['main bus amps', 'bus amps', 'main amps', 'bus rating'],
                    'main_breaker': ['main circuit breaker', 'main breaker', 'mcb'],
                    'mounting': ['mounting', 'mount'],
                    'feed': ['feed from', 'fed from', 'feed'],
                    'location': ['location', 'loc'],
                }
                
                # Check both sets of label/value columns: (A,B) and (N,O)
                label_value_columns = [(1, 2), (14, 15)]  # (A,B) and (N,O)
                
                for label_col, value_col in label_value_columns:
                    # Read labels from rows 2-9 and populate values
                    for row_num in range(2, 10):
                        label_cell = ws.cell(row=row_num, column=label_col)
                        if label_cell.value:
                            label = str(label_cell.value).lower().strip().rstrip(':')
                            
                            # Find matching spec key
                            matched_value = None
                            for spec_key, keywords in label_mapping.items():
                                if any(keyword in label for keyword in keywords):
                                    # Check if we have the spec value, otherwise mark as MISSING
                                    matched_value = panel_specs.get(spec_key, "MISSING")
                                    break
                            
                            # If no keyword matched, default to MISSING for any labeled field
                            if matched_value is None:
                                matched_value = "MISSING"
                                logger.info(f"No keyword match for label '{label}' - marking as MISSING")
                            
                            # Update value cell with matched value or MISSING
                            col_letter = openpyxl.utils.get_column_letter(value_col)
                            ws.cell(row=row_num, column=value_col, value=matched_value)
                            logger.info(f"Updated {label} ({col_letter}{row_num}) to: {matched_value}")
            
            # Find the data table header block (may span multiple rows)
            # Look for keywords like "LOAD DESCRIPTION", "CKT", etc.
            first_header_row = None
            last_header_row = None
            
            for row_num in range(1, min(20, ws.max_row + 1)):
                has_header_keyword = False
                for col_num in range(1, min(16, ws.max_column + 1)):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower()
                        if any(keyword in cell_lower for keyword in ['load description', 'ckt', 'breaker', 'phase', 'amp', 'pole']):
                            has_header_keyword = True
                            if first_header_row is None:
                                first_header_row = row_num
                            last_header_row = row_num
                            break
                
                # If we found headers and now hit a row without header keywords, stop
                if first_header_row and not has_header_keyword:
                    break
            
            # Data starts after the last header row
            if last_header_row:
                data_start_row = last_header_row + 1
                logger.info(f"Found header block from row {first_header_row} to {last_header_row}, data starts at row {data_start_row}")
            else:
                data_start_row = 12
                logger.warning(f"Could not find data table header, using default row {data_start_row}")
            
            # Fill in circuit data starting from data_start_row
            # Find which columns contain circuit number and description
            circuit_col = None
            desc_col = None
            
            # Check header row for circuit and description columns
            header_row = data_start_row - 1
            for col_num in range(1, min(16, ws.max_column + 1)):
                cell_value = ws.cell(row=header_row, column=col_num).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower()
                    if 'ckt' in cell_lower or ('circuit' in cell_lower and 'breaker' not in cell_lower):
                        circuit_col = col_num
                    elif 'load' in cell_lower or 'desc' in cell_lower:
                        desc_col = col_num
            
            # Default columns if not found: circuit in col 7, description in col 1
            if not circuit_col:
                circuit_col = 7
            if not desc_col:
                desc_col = 1
            
            logger.info(f"Filling circuit data: circuit col={circuit_col}, description col={desc_col}")
            
            # Fill circuit data
            for idx, (circuit_num, description) in enumerate(circuits):
                row_num = data_start_row + idx
                ws.cell(row=row_num, column=circuit_col, value=circuit_num)
                ws.cell(row=row_num, column=desc_col, value=description)
            
            logger.info(f"Applied template with {len(circuits)} circuits")
            
        except Exception as e:
            logger.warning(f"Failed to apply template {template_path.name}: {e.__class__.__name__}: {e}. Falling back to basic format.")
            # Fall through to basic format creation below
            template_path = None
    
    if not template_path:
        # Create basic formatted schedule
        logger.info("No template found, creating basic formatted schedule")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Panel_Schedule'
        
        # Create headers with basic formatting
        headers = ['Panel', 'Circuit', 'Description']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        
        # Fill in circuit data
        for row_idx, (circuit_num, description) in enumerate(circuits, 2):
            ws.cell(row=row_idx, column=1, value=panel_name)
            ws.cell(row=row_idx, column=2, value=circuit_num)
            ws.cell(row=row_idx, column=3, value=description)
    
    wb.save(output_path)
    logger.info(f"Saved panel schedule to {output_path}")
    return output_path
