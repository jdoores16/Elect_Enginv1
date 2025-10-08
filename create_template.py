#!/usr/bin/env python3
"""
Create a professional panelboard schedule template
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Panel Schedule"

# Define colors
header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:I1')
ws['A1'] = 'PANELBOARD SCHEDULE'
ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws['A1'].fill = header_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Panel Info Section
ws.merge_cells('A2:C2')
ws['A2'] = 'Panel ID:'
ws['A2'].font = Font(bold=True)
ws['D2'] = ''  # User fills this

ws.merge_cells('E2:F2')
ws['E2'] = 'Voltage:'
ws['E2'].font = Font(bold=True)
ws['G2'] = '208Y/120V'

ws.merge_cells('H2:H2')
ws['H2'] = 'Main:'
ws['H2'].font = Font(bold=True)
ws['I2'] = '225A'

# Column Headers
headers = ['Circuit', 'Description', 'Breaker\n(A)', 'Poles', 'Load\n(VA)', 'Phase\nA', 'Phase\nB', 'Phase\nC', 'Notes']
col_widths = [10, 35, 10, 8, 12, 8, 8, 8, 20]

for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=4, column=col_num, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_thin
    ws.column_dimensions[get_column_letter(col_num)].width = width

ws.row_dimensions[4].height = 35

# Add sample rows with formatting
for row in range(5, 45):  # 40 circuit rows
    for col in range(1, 10):
        cell = ws.cell(row=row, column=col)
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center' if col in [1,3,4,6,7,8] else 'left', vertical='center')
        
        # Add circuit numbers for odd rows
        if col == 1 and row % 2 == 1:
            cell.value = (row - 4) // 2 + 1

# Totals row
total_row = 45
ws.merge_cells(f'A{total_row}:B{total_row}')
ws[f'A{total_row}'] = 'TOTAL CONNECTED LOAD'
ws[f'A{total_row}'].font = Font(bold=True)
ws[f'A{total_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws[f'E{total_row}'] = f'=SUM(E5:E44)'
ws[f'E{total_row}'].font = Font(bold=True)
ws[f'E{total_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws[f'E{total_row}'].border = border_thin

# Notes section
notes_row = 47
ws.merge_cells(f'A{notes_row}:I{notes_row}')
ws[f'A{notes_row}'] = 'Notes: All circuits sized per NEC Article 220. Verify actual loads before installation.'
ws[f'A{notes_row}'].font = Font(italic=True, size=9)
ws[f'A{notes_row}'].alignment = Alignment(horizontal='left')

# Save
wb.save('templates/panelboard_schedule_template.xlsx')
print("✓ Template created: templates/panelboard_schedule_template.xlsx")
